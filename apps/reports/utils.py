import csv
import io
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def export_csv(records, filename='attendance_report.csv'):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Student', 'Course', 'Date', 'Time', 'Status'])
    for r in records:
        writer.writerow([
            r.student.get_full_name() or r.student.username,
            r.course.code,
            r.session.session_date,
            r.scan_time.strftime('%H:%M:%S'),
            r.status,
        ])
    return output.getvalue()


def export_pdf(records, title='Attendance Report'):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles['Title']), Spacer(1, 20)]
    data = [['Student', 'Course', 'Date', 'Time', 'Status']]
    for r in records:
        data.append([
            r.student.get_full_name() or r.student.username,
            r.course.code,
            str(r.session.session_date),
            r.scan_time.strftime('%H:%M:%S'),
            r.status,
        ])
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    elements.append(table)
    doc.build(elements)
    return buffer.getvalue()


def export_excel(records, title='Attendance Report'):
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='2563eb', end_color='2563eb', fill_type='solid')
    headers = ['Student', 'Course', 'Date', 'Time', 'Status']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    for row, r in enumerate(records, 2):
        ws.cell(row=row, column=1, value=r.student.get_full_name() or r.student.username)
        ws.cell(row=row, column=2, value=r.course.code)
        ws.cell(row=row, column=3, value=str(r.session.session_date))
        ws.cell(row=row, column=4, value=r.scan_time.strftime('%H:%M:%S'))
        ws.cell(row=row, column=5, value=r.status)
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 20
    wb.save(output)
    return output.getvalue()
